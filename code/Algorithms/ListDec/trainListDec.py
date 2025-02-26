from Algorithms.ListDec.globalListDec import gListDec
from Algorithms.ListDec.localListDec import lListDec
from Algorithms.ListDec.nodeListDec import nListDec
from federatedFrameW.ftrain.trainCentralizedFL import CentralizedFL
from federatedFrameW.utils.model_utils import general_CenFL_filename
from federatedFrameW.utils.data_utils import read_data, read_user_data_vad

import copy
import h5py
from openpyxl import Workbook
import time
import numpy as np


class ListDec(CentralizedFL):
    '''
    Federated Learning with Byzantine Attack and List Decodable robust scheme

    kwargs:
        - dataset: the name of the dataset
        - device: the device to train the model on
        - model: the origin model for deepcopy
        - name: the name of the algorithm
        - lImp: the local implementation
        - gImp: the global implementation
        - nImp: the node implementation
        - hyperparams: the hyperparameters
            - model_name: the name of the model
            - batch_size: batch size
            - total_epochs: total number of epochs
            - local_epochs: int, number of epochs for local training
            - beta: global momentum
            - num_aggregate_locals: number of local models to aggregate
            - learning_rate: learning rate
            - times: the number of times to repeat the experiment
            - loss_name: the name of the loss function
            - optimizer_name: the name of the optimizer
    '''

    def __init__(self, *args, **kwargs):
        kwargs['lImp'] = lListDec
        kwargs['gImp'] = gListDec
        kwargs['nImp'] = nListDec
        super().__init__(*args, **kwargs)

    def gen_locals(self):
        '''Override, add validation set'''
        data = read_data(dataset=self.dataset, data_distrb=self.local_data_distrb)
        self.total_users = len(data[0])
        print("trainer.total_users: %d" % (self.total_users))
        for i in range(self.total_users):
            id, train_data, test_data, valid_data = read_user_data_vad(i, data, self.dataset, self.hyperparams)
            model = copy.deepcopy(self.model)
            hyperparams = copy.deepcopy(self.hyperparams)
            local = self.lImp(id=id, device=self.device, model=model, train_data=train_data, test_data=test_data,valid_data=valid_data,
                              hyperparams=hyperparams)
            self.flocals.append(local)
            self.total_train_samples += local.train_samples

    def res_file_name(self, tag=''):
        return general_CenFL_filename(self, tag) + "_" + str(self.at) + "_" + str(self.corp_rate) + "_" + str(self.times)

    def train(self):
        self.pre_train()
        for action_fn in self.actions:
            self.pre_action(action_fn)
            action_fn()
            self.post_action(action_fn)
            if self.c == -1:
                break
        self.post_train()

    def eval_global_model(self):
        for fn in self.fnodes:
            self.c = fn.eval_global()
            if self.c == 1 or self.c == 2 or self.c == 5 or ((self.c)%self.eval_every == 0) or self.c == -1:
                print("--------------Evaluate Global Model time: {:.04f}".format(time.time() - self.t), " ----------------")

    def pre_local_train(self):
        self.t = time.time()
        self.c += 1
        if self.c == 1 or self.c == 2 or self.c == 5 or ((self.c)%self.eval_every == 0):
            print("-------------Round number: {0:05d}".format(self.c), " -------------")
        for fn in self.fnodes:
            fn.sample_local() # sample local to train

    def train_local_model(self):
        for fn in self.fnodes:
            fn.train(round=self.c)
        if self.c == 1 or self.c == 2 or self.c == 5 or ((self.c)%self.eval_every == 0):
            print("--------------Local Train time: {:.04f}".format(time.time() - self.t), " ----------------")

    def save_global_results(self):
        '''save results, override'''
        file_name = self.res_file_name(tag='_g')
        self.save_xls()
        
        # save results as a .h5 file
        for fn in self.fnodes:
            for fg in fn.fglobals:
                with h5py.File(self.log_fp+'.h5', 'w') as hf:
                    print('save H5 log: ', self.log_fp+'.h5')
                    hf.create_dataset('rs_glob_acc', data=fg.rs_glob_acc) # every eval_every round
                    # hf.create_dataset('rs_train_acc', data=fg.rs_train_acc)
                    hf.create_dataset('clean_tl', data=fg.clean_tl) # for clean only
                    hf.create_dataset('rs_valid_loss', data=fg.min_valid_loss) # each round
                    hf.create_dataset('rs_test_loss', data=fg.rs_test_loss) # every eval_every round
                    hf.create_dataset('global_paras_update', data=fg.para_group_update)
                    hf.create_dataset('votes', data=fg.votes) # each round
                    hf.create_dataset('avg_hlocal_valid_ls', data=fg.avg_hlocal_vls) # each round
                    # hf.create_dataset('paras_ts', data = fg.paras_ts) # self.paras_ts
                    
                    # hf.create_dataset('vote_stats_t', data = list(fg.vote_stats.keys()))
                    # hf.create_dataset('loss_gap_mean', data = [l[0] for l in fg.vote_stats.values()]) # {ground: {mean, std, true/false}}
                    # hf.create_dataset('loss_gap_std', data = [l[1] for l in fg.vote_stats.values()])
                    # hf.create_dataset('loss_rate_mean', data = [l[2] for l in fg.vote_stats.values()]) # {ground: {mean, std, true/false}}
                    # hf.create_dataset('loss_rate_std', data = [l[3] for l in fg.vote_stats.values()])
                    # hf.create_dataset('loss_drop_t', data = list(fg.num_no_loss_drop_dic.keys()))
                    # hf.create_dataset('num_no_loss_drop', data = list(fg.num_no_loss_drop_dic.values()))  

                    hf.create_dataset('vote_right_t', data = list(fg.vote_right_rate.keys()))
                    vote_right_rate = [1 if v is True else 0 for v in fg.vote_right_rate.values()]
                    vote_right_rate.append(sum(vote_right_rate) / (len(vote_right_rate) + 1e-9)) 
                    hf.create_dataset('vote_right_rate', data = vote_right_rate)

                    hf.create_dataset('vote_false_t', data = list(fg.vote_false_rate.keys()))
                    vote_false_rate = [1 if v is True else 0 for v in fg.vote_false_rate.values()]
                    vote_false_rate.append(sum(vote_false_rate) / (len(vote_false_rate) + 1e-9))
                    hf.create_dataset('vote_false_rate', data = vote_false_rate)

                    hf.create_dataset('train_loss_t', data = list(fg.cor_tl.keys()))
                    hf.create_dataset('train_loss', data = list(fg.cor_tl.values()))
                    print(f'train_loss_t[:10]: {list(fg.cor_tl.keys())[:10]}')
                    print(f'train_loss[:10]: {list(fg.cor_tl.values())[:10]}')
                    hf.close()

                    # self.save_aim("results/"+file_name, train_acc=fg.rs_glob_acc, train_loss=fg.min_valid_loss,
                    #          glob_acc=fg.rs_glob_acc,
                    #          test_loss=fg.rs_test_loss)
    
    def save_xls(self):
        print('save_xls: ', self.log_fp+'.xlsx')
        book = Workbook() # openpyxl.load_workbook(self.log_fp+'.xlsx') # Workbook()
        sheet = book.active
        row = 1
        col = 1
        for fn in self.fnodes:
            for fg in fn.fglobals:
                # header
                # record test results
                sheet.cell(row=row , column=col).value = 'test_acc'
                col += 1
                sheet.cell(row=row , column=col).value = 'test_loss'
                col += 1
                # record training results
                l1 = len(fg.avg_hlocal_vls[0])
                lt = len(fg.paras_ts[0])
                l2 = len(fg.votes[0])
                # l3 = len(fg.)
                sheet.cell(row=row, column=col).value = 'gmdls_update' # fg.para_group_update
                col += 1
                sheet.cell(row=row, column=col).value = 'broadcast_gmdl' # fg.brad_gmdl
                col += 1
                sheet.cell(row=row, column=col).value = 'avg_hlocal_valid_ls' # fg.avg_hlocal_tls
                col += l1
                sheet.cell(row=row, column=col).value = 'paras_ts' # fg.paras_ts
                col += lt
                sheet.cell(row=row, column=col).value = 'votes' # fg.votes
                col += l2
                sheet.cell(row=row, column=col).value = 'best_train_loss' # fg.min_valid_loss
                col += 1
                sheet.cell(row=row, column=col).value = 'loss_up' # if sampled hlocal when loss increase
                
                # data
                row += 1
                col = 1
                idx = 0
                for (brad_gmdl,x,t,y,z) in zip(fg.brad_gmdl, fg.avg_hlocal_vls, fg.paras_ts, fg.votes, fg.min_valid_loss):
                    if (row-1) == 1 or (row-1) == 2 or (row-1) == 5 or ((row-1)%self.eval_every ==0):
                        sheet.cell(row=row , column=col).value = fg.rs_glob_acc[idx]
                        sheet.cell(row=row , column=col+1).value = fg.rs_test_loss[idx]
                        idx += 1
                    col += 2

                    sheet.cell(row=row, column=col).value = (row-1) in fg.para_group_update
                    col += 1
                    sheet.cell(row=row, column=col).value = brad_gmdl
                    col += 1

                    for i in range(l1):
                        sheet.cell(row=row, column=col+i).value = x[i]
                    col += l1
                    for i in range(lt):
                        sheet.cell(row=row, column=col+i).value = t[i]
                    col += lt
                    for i in range(l2):
                        sheet.cell(row=row, column=col+i).value = y[i]
                    col += l2
                    sheet.cell(row=row, column=col).value = z
                    
                    if ( (row-1) in fg.loss_up.keys()):
                        sheet.cell(row=row, column=col+1).value = fg.loss_up[row-1]
    
                    row += 1
                    col = 1
                # if training stop at some round which can not divided by eval_every(pre-stop)
                if idx <= len(fg.rs_glob_acc)-1:
                    sheet.cell(row=row-1 , column=col).value = fg.rs_glob_acc[-1]
                    sheet.cell(row=row-1 , column=col+1).value = fg.rs_test_loss[-1]

        book.save(self.log_fp+'.xlsx')